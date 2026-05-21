"""BMC export — JSON / XLSX / PDF.

Takes a serialized ``BMCRead`` (decoupled from the ORM) and renders a
downloadable artifact. All pure-Python (openpyxl, fpdf2) — no Pandoc/LaTeX
system dependency, so it runs anywhere the backend runs.

Every format includes the SOURCES so the export is self-documenting — an
analyst can hand the PDF/XLSX to their IC and every claim traces to a filing
page. That provenance is the whole point of PRISM's BMC.
"""

from __future__ import annotations

import io
import re

from src.schemas.bmc import BMCRead

# (media_type, file extension) per format.
EXPORT_FORMATS: dict[str, tuple[str, str]] = {
    "json": ("application/json", "json"),
    "xlsx": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"),
    "pdf": ("application/pdf", "pdf"),
}


def export_bmc(bmc: BMCRead, fmt: str) -> bytes:
    if fmt == "json":
        return _to_json(bmc)
    if fmt == "xlsx":
        return _to_xlsx(bmc)
    if fmt == "pdf":
        return _to_pdf(bmc)
    raise ValueError(f"Unsupported export format {fmt!r}. Valid: {list(EXPORT_FORMATS)}")


def filename_for(bmc: BMCRead, fmt: str) -> str:
    ext = EXPORT_FORMATS[fmt][1]
    return f"{bmc.ticker}_BMC_v{bmc.version}.{ext}"


# ── JSON ────────────────────────────────────────────────────────────────────


def _to_json(bmc: BMCRead) -> bytes:
    return bmc.model_dump_json(indent=2).encode("utf-8")


# ── XLSX ────────────────────────────────────────────────────────────────────


def _to_xlsx(bmc: BMCRead) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()

    # Sheet 1 — Canvas overview.
    ws = wb.active
    ws.title = "Canvas"
    ws.append([f"Business Model Canvas — {bmc.ticker}"])
    ws["A1"].font = Font(size=14, bold=True)
    ws.append([f"Version {bmc.version}", bmc.fiscal_period or "", f"status: {bmc.status}"])
    if bmc.overall_confidence is not None:
        ws.append([f"Overall confidence: {round(bmc.overall_confidence * 100)}%"])
    ws.append([])
    ws.append(["Block", "Confidence", "Status", "Bullets"])
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
    for b in bmc.blocks:
        ws.append(
            [
                b.title,
                f"{round(b.confidence * 100)}%" if b.status == "ok" else "—",
                b.status,
                "\n".join(b.summary_bullets) if b.summary_bullets else "(no evidence)",
            ]
        )
    # Wrap the bullets column.
    for row in ws.iter_rows(min_row=header_row + 1, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["D"].width = 80

    # Sheet 2 — Sources / audit trail.
    audit = wb.create_sheet("Sources")
    audit.append(["Block", "Marker", "Page", "Excerpt"])
    for cell in audit[1]:
        cell.font = Font(bold=True)
    for b in bmc.blocks:
        for ev in b.evidence:
            audit.append([b.title, ev.marker, ev.page_number, ev.excerpt])
    audit.column_dimensions["A"].width = 24
    audit.column_dimensions["D"].width = 100
    for row in audit.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Sheet 3 — Contradictions (only if any).
    if bmc.contradictions:
        con = wb.create_sheet("Contradictions")
        con.append(["Block A", "Block B", "Issue"])
        for cell in con[1]:
            cell.font = Font(bold=True)
        for c in bmc.contradictions:
            con.append([c.block_a, c.block_b, c.issue])
        con.column_dimensions["C"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF ─────────────────────────────────────────────────────────────────────


def _pdf_safe(text: str) -> str:
    """fpdf2's core fonts are latin-1. Replace common non-latin glyphs (₹, —,
    smart quotes, bullets) so export never crashes on Indian filing text."""
    repl = {
        "₹": "Rs.",  # ₹
        "—": "-", "–": "-",  # em/en dash
        "‘": "'", "’": "'",
        "“": '"', "”": '"',
        "•": "-", "→": "->",
        "≥": ">=", "≤": "<=",
    }
    for k, v in repl.items():
        text = text.replace(k, v)
    # Drop anything still outside latin-1.
    return text.encode("latin-1", "replace").decode("latin-1")


def _to_pdf(bmc: BMCRead) -> bytes:
    from fpdf import FPDF

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, _pdf_safe(f"Business Model Canvas — {bmc.ticker}"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    meta = f"Version {bmc.version}"
    if bmc.fiscal_period:
        meta += f"  |  {bmc.fiscal_period}"
    meta += f"  |  status: {bmc.status}"
    if bmc.overall_confidence is not None:
        meta += f"  |  confidence: {round(bmc.overall_confidence * 100)}%"
    pdf.set_text_color(110, 119, 133)
    pdf.cell(0, 6, _pdf_safe(meta), new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 6, _pdf_safe("Every claim is cited to a primary filing — see Sources."), new_x="LMARGIN", new_y="NEXT")
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    for b in bmc.blocks:
        pdf.set_font("Helvetica", "B", 12)
        conf = f"  ({round(b.confidence * 100)}%)" if b.status == "ok" else "  (no evidence)"
        pdf.cell(0, 8, _pdf_safe(b.title + conf), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        if b.summary_bullets:
            for bullet in b.summary_bullets:
                pdf.multi_cell(0, 5, _pdf_safe("- " + bullet))
        else:
            pdf.set_text_color(150, 150, 150)
            pdf.multi_cell(0, 5, _pdf_safe("No filing evidence for this block."))
            pdf.set_text_color(0, 0, 0)
        pdf.ln(2)

    if bmc.contradictions:
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 8, _pdf_safe("Cross-block inconsistencies"), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        for c in bmc.contradictions:
            pdf.multi_cell(0, 5, _pdf_safe(f"- {c.block_a} <-> {c.block_b}: {c.issue}"))

    # Sources appendix.
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, _pdf_safe("Sources"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    for b in bmc.blocks:
        for ev in b.evidence:
            page = f"p.{ev.page_number}" if ev.page_number is not None else "filing"
            head = f"[{b.title} {ev.marker} — {page}]"
            pdf.set_font("Helvetica", "B", 9)
            pdf.multi_cell(0, 4.5, _pdf_safe(head))
            pdf.set_font("Helvetica", "", 9)
            # Trim long table dumps for the appendix.
            excerpt = re.sub(r"\s+", " ", ev.excerpt)[:400]
            pdf.multi_cell(0, 4.5, _pdf_safe(excerpt))
            pdf.ln(1)

    out = pdf.output()
    return bytes(out)
